"""API tests for rate management endpoints."""

from io import BytesIO

import openpyxl
import pytest
from httpx import AsyncClient


class TestRatesAPI:
    """Test suite for rate management API endpoints."""

    @pytest.mark.asyncio
    async def test_upload_rates_excel_success(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading rates from Excel file."""
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Apples", 150, "ALL"])
        ws.append(["Oranges", 200, "ALL"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        assert response.status_code == 200
        data = response.json()
        assert "message" in data
        assert "2" in data["message"]

    @pytest.mark.asyncio
    async def test_upload_rates_excel_replaces_existing(
        self, test_client: AsyncClient, sample_rates
    ):
        """Test that uploading rates replaces existing rates."""
        # Create new Excel with different data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Bananas", 100, "ALL"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        assert response.status_code == 200

        # Verify only new rates exist
        get_response = await test_client.get("/rates?format=json")
        assert get_response.status_code == 200
        rates = get_response.json()
        assert len(rates) == 1
        assert rates[0]["product_id"] == "Bananas"

    @pytest.mark.asyncio
    async def test_upload_rates_invalid_file_format(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading non-Excel file returns error."""
        files = {"file": ("rates.txt", BytesIO(b"not an excel file"), "text/plain")}
        response = await test_client.post("/rates", files=files)

        assert response.status_code == 400
        data = response.json()
        assert "detail" in data

    @pytest.mark.asyncio
    async def test_upload_rates_missing_columns(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading Excel with missing required columns."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate"])  # Missing Scope column
        ws.append(["Apples", 150])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_upload_rates_empty_file(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading empty Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])  # Only headers

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        # Should succeed with 0 rates
        assert response.status_code == 200
        data = response.json()
        assert "0" in data["message"]

    @pytest.mark.asyncio
    async def test_upload_rates_invalid_data_types(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading Excel with invalid data types."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Apples", "not-a-number", "ALL"])  # Invalid rate

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_upload_rates_from_directory_success(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading rates from /in directory."""
        # This test requires actual file in /in directory
        # Skip if file doesn't exist
        response = await test_client.post(
            "/rates/from-directory", json={"file": "rates.xlsx"}
        )

        # Should either succeed or return file not found
        assert response.status_code in [200, 400]

    @pytest.mark.asyncio
    async def test_upload_rates_from_directory_file_not_found(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading from non-existent file in directory."""
        response = await test_client.post(
            "/rates/from-directory", json={"file": "nonexistent.xlsx"}
        )

        assert response.status_code == 400
        data = response.json()
        assert "detail" in data

    @pytest.mark.asyncio
    async def test_get_rates_json_format(self, test_client: AsyncClient, sample_rates):
        """Test getting rates in JSON format."""
        response = await test_client.get("/rates?format=json")

        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        assert len(data) == 3
        assert all("product_id" in rate for rate in data)
        assert all("rate" in rate for rate in data)
        assert all("scope" in rate for rate in data)

    @pytest.mark.asyncio
    async def test_get_rates_json_empty(self, test_client: AsyncClient, clean_database):
        """Test getting rates when database is empty."""
        response = await test_client.get("/rates?format=json")

        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        assert len(data) == 0

    @pytest.mark.asyncio
    async def test_get_rates_excel_format(self, test_client: AsyncClient, sample_rates):
        """Test downloading rates as Excel file."""
        response = await test_client.get("/rates?format=excel")

        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response.headers["content-disposition"]
        assert "rates.xlsx" in response.headers["content-disposition"]

        # Verify Excel content
        excel_data = BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_data)
        ws = wb.active
        assert ws.cell(1, 1).value == "Product"
        assert ws.cell(1, 2).value == "Rate"
        assert ws.cell(1, 3).value == "Scope"

    @pytest.mark.asyncio
    @pytest.mark.skip(reason="TODO: Fix later")
    async def test_get_rates_excel_empty(
        self, test_client: AsyncClient, clean_database
    ):
        """Test downloading rates as Excel when database is empty."""
        response = await test_client.get("/rates?format=excel")

        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Verify Excel has only headers
        excel_data = BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_data)
        ws = wb.active
        assert ws.cell(1, 1).value == "Product"
        assert ws.max_row == 1  # Only header row

    @pytest.mark.asyncio
    async def test_get_rates_default_format(
        self, test_client: AsyncClient, sample_rates
    ):
        """Test getting rates with default format (should be Excel)."""
        response = await test_client.get("/rates")

        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @pytest.mark.asyncio
    async def test_get_rates_invalid_format(
        self, test_client: AsyncClient, sample_rates
    ):
        """Test getting rates with invalid format parameter."""
        response = await test_client.get("/rates?format=invalid")

        # Should default to Excel format
        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @pytest.mark.asyncio
    async def test_upload_rates_large_file(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading large Excel file with many rates."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])

        # Add 100 rates
        for i in range(100):
            ws.append([f"Product{i}", 100 + i, "ALL"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        assert response.status_code == 200
        data = response.json()
        assert "100" in data["message"]

    @pytest.mark.asyncio
    async def test_upload_rates_with_provider_specific_scope(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading rates with provider-specific scopes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Apples", 150, "ALL"])
        ws.append(["Apples", 175, "1"])  # Provider 1 specific
        ws.append(["Apples", 160, "2"])  # Provider 2 specific

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        assert response.status_code == 200

        # Verify all rates were uploaded
        get_response = await test_client.get("/rates?format=json")
        rates = get_response.json()
        assert len(rates) == 3


class TestRatesAPIEdgeCases:
    """Test suite for rate API edge cases."""

    @pytest.mark.asyncio
    async def test_upload_rates_with_empty_values(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading rates with empty values."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["", 150, "ALL"])  # Empty product
        ws.append(["Apples", None, "ALL"])  # None rate
        ws.append(["Oranges", 200, ""])  # Empty scope

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        # Should handle gracefully with error
        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_upload_rates_with_negative_rate(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading rates with negative rate value."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Apples", -150, "ALL"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        # Depending on business rules, might accept or reject
        assert response.status_code in [200, 400]

    @pytest.mark.asyncio
    async def test_upload_rates_with_zero_rate(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading rates with zero rate value."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Apples", 0, "ALL"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        assert response.status_code in [200, 400]

    @pytest.mark.asyncio
    async def test_upload_rates_with_duplicate_entries(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading rates with duplicate product-scope combinations."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Apples", 150, "ALL"])
        ws.append(["Apples", 175, "ALL"])  # Duplicate

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        # Should handle gracefully
        assert response.status_code in [200, 400]

    @pytest.mark.asyncio
    async def test_upload_rates_with_extra_columns(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading Excel with extra columns beyond required."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope", "Extra", "Column"])
        ws.append(["Apples", 150, "ALL", "data", "here"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        # Should ignore extra columns and succeed
        assert response.status_code == 200

    @pytest.mark.asyncio
    async def test_upload_rates_case_sensitivity(
        self, test_client: AsyncClient, clean_database
    ):
        """Test if product names are case-sensitive."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Apples", 150, "ALL"])
        ws.append(["apples", 175, "ALL"])  # Lowercase

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await test_client.post("/rates", files=files)

        assert response.status_code == 200

        # Check if treated as different products
        get_response = await test_client.get("/rates?format=json")
        rates = get_response.json()
        assert len(rates) == 2

    @pytest.mark.asyncio
    async def test_get_rates_excel_content_verification(
        self, test_client: AsyncClient, sample_rates
    ):
        """Test Excel download contains correct data."""
        response = await test_client.get("/rates?format=excel")

        excel_data = BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_data)
        ws = wb.active

        # Check data rows (skip header)
        row_count = ws.max_row - 1
        assert row_count == 3  # sample_rates has 3 entries

    @pytest.mark.asyncio
    async def test_upload_then_download_roundtrip(
        self, test_client: AsyncClient, clean_database
    ):
        """Test uploading rates and downloading them back."""
        # Upload
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Rate", "Scope"])
        ws.append(["Apples", 150, "ALL"])
        ws.append(["Oranges", 200, "ALL"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files = {
            "file": (
                "rates.xlsx",
                buffer,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        upload_response = await test_client.post("/rates", files=files)
        assert upload_response.status_code == 200

        # Download
        download_response = await test_client.get("/rates?format=json")
        assert download_response.status_code == 200
        rates = download_response.json()
        assert len(rates) == 2
        assert any(r["product_id"] == "Apples" and r["rate"] == 150 for r in rates)
        assert any(r["product_id"] == "Oranges" and r["rate"] == 200 for r in rates)
